"""Compiles vendor selection scorecards into results by vendor.
IMPORTANT: There must be consistent worksheet (tab) and column naming:
Sheets must be either 'Ind. Scorecard' or 'Cons. Scorecard', while
columns must include ones named: 'ID', 'Cat+SubCat', 'Expertise',
and a column for each vendor score."""

import openpyxl, pprint
from os import system
from openpyxl.utils import get_column_letter, column_index_from_string


def get_base_name():
	"""Get base name for all scorecards"""

	base_name = input('Enter base name for scorecard workbooks '
		'(i.e., name without the numeral at the end): ')
	print("Enter the total number of scorecards")
	num_of_scorecards = input("(if this is wrong, the "
		"program either won't work or scorecards will be "
		"missed!): ")

	num_of_scorecards = int(num_of_scorecards)

	return base_name, num_of_scorecards


def get_weighting_name():
	"""Get name for weighting sheet from user"""

	weighting_name = input('Enter name of weighting sheet '
		'workbook (without extension): ')

	return weighting_name


def choose_scorecard_type():
	"""User selects 'Ind.'' or 'Cons' scorecard"""

	while True:
		scorecard_type = input("Enter '1' for Ind. Scorecard or '2' "
			"for Cons. Scorecard: ")
		if scorecard_type == '1':
			sheet_name = 'Ind. Scorecard'
			break
		elif scorecard_type == '2':
			sheet_name = 'Cons. Scorecard'
			break
		else:
			print("You must enter 1 or 2!")
			continue

	return sheet_name


def get_user_data():
	"""User inputs a base name for the scorecard workbooks so Python
	can open them all by iterating over a range of numbers (provided
	by user as total number of workbooks. User must also tell
	Python whether to look for the 'Ind.' or 'Cons.' Scorecard."""
	system('cls')
	print("Scorcard compilation program.")
	print("To begin, ensure that all scorecards and the weighting "
		"sheet are in the same folder")
	print("as this program itself. Then, rename all scorecard "
		"workbooks so they have the same")
	print("base name, followed by a space and an index in () starting with (1). "
		"Example: Client Scorecard (1),")
	print("Client Scorecard (2), etc. (Windows bulk renaming will do this automatically.)")
	input("When this is done, press enter to begin.")

	base_name, num_of_scorecards = get_base_name()

	sheet_name = choose_scorecard_type()	

	weighting_name = get_weighting_name()

	print("Please ensure that all Excel files are closed.")
	input("When ready, press enter.")

	return base_name, num_of_scorecards, sheet_name, weighting_name


def get_vendors(base_name, sheet_name):
	"""Opens the first workbook and searches for 'Score' to find
	vendor names"""

	while True:
		print(f'Opening {base_name} (1) to look for vendors...')
		try:
			wb = openpyxl.load_workbook(f'{base_name} (1).xlsx', data_only=True)
		except FileNotFoundError:
			print(f"File {base_name} (1).xlsx NOT FOUND!")
			print("Double check that it's in the same folder as "
				"scorecard_compilation.py,")
			input("then press 'enter' to re-enter the base name.")
			base_name, num_of_scorecards = get_base_name()
			print()
			continue
		else:
			break

	while True:
		try:
			sheet = wb[sheet_name]
		except:
			print(f"{sheet_name} NOT FOUND! Wrong scorecard type?")
			input("Press 'enter' try re-enter scorecard type.")
			sheet_name = choose_scorecard_type()
			print()
			continue
		else:
			break

	#Go across column headers (row 3) and search for "[vendor] Score"
	#to find vendors
	vendors = []
	for column in range(1, sheet.max_column + 1):
		column_letter = get_column_letter(column)
		if sheet[f'{column_letter}3'].value is not None:
			if "Score" in sheet[f'{column_letter}3'].value:
				vendor = str(sheet[f'{column_letter}3'].value)
				#Remove " Score" from string, leaving only vendor name
				vendor = vendor.replace(" Score", "")
				vendors.append(vendor)

	num_of_vendors = len(vendors)

	print(f"Number of vendors: {num_of_vendors}")
	print("Vendors: ", end="")
	print(*vendors, sep=', ')
	correct = input("Is this correct (y/n)? ")
	if correct == 'n':
		exit()

	#Python can't use dictionaries with spaces in their names
	nsv = []
	for vendor in vendors:
		no_space_vendor = vendor.replace(' ', '_')
		nsv.append(no_space_vendor)

	return vendors, nsv, num_of_vendors, base_name, sheet_name


def create_aggregate_file(base_name, num_of_scorecards, sheet_name,
	vendors, nsv, num_of_vendors):
	"""Opens the scorecards and stores their contents in a Python
	dictionary."""

	aggregate_dict = {}
	
	for i in range(num_of_scorecards):
		while True:
			print(f'Opening {base_name} ({i+1})...')
			try:
				wb = openpyxl.load_workbook(f'{base_name} ({i+1}).xlsx',
					data_only=True)
			except FileNotFoundError:
				print(f"{base_name} ({i+1}) NOT FOUND! Please check that all"
					" files have the same base name and are numbered in "
					"order")
				input("Press 'enter' to confirm and try again.")
				print()
				continue
			else:
				break

		sheet = wb[sheet_name]

		#Set up the columns you need when you open first workbook
		if i == 0:
			#ID, Cat+SubCat, Expertise are always needed
			standard_columns = ['A', 'B', 'F']
			vendor_score_columns = []
			#Find column indices and convert to letters for each
			#column containing '[Vendor] Score'
			for vendor in vendors:
				for column in range(1, sheet.max_column + 1):
					column_letter = get_column_letter(column)
					if sheet[f'{column_letter}3'].value == f"{vendor} Score":
						vendor_score_columns.append(column_letter)

			headers = []
			for column in standard_columns:
				headers.append(sheet[f'{column}3'].value)
			for column in vendor_score_columns:
				headers.append(sheet[f'{column}3'].value)

		for row in range(4, sheet.max_row + 1):
			for x in range (0, num_of_vendors):
				vendor_name = nsv[x]
				item_id = sheet[standard_columns[0] + str(row)].value
				cat_plus_subcat = sheet[standard_columns[1] + str(row)].value
				#Replace en and em dashes with hyphens:
				if cat_plus_subcat is not None:
					cat_plus_subcat = cat_plus_subcat.replace(u"\u2013", "-")
					cat_plus_subcat = cat_plus_subcat.replace(u"\u2014", "-")
				expertise = sheet[standard_columns[2] + str(row)].value
				expertise = int(0 if expertise is None else expertise)
				score = sheet[vendor_score_columns[x] + str(row)].value
				if score is None or score == '' or score == ' ':
					score = 0
				#print(cat_plus_subcat,score)
				score = int(score)

				#Any blank scores are not counted--and their expertise
				#values must not be counted, either!
				if score == 0:
					expertise = 0
				weighted_score = expertise * score

				#If the vendor name key hasn't been seen before,
				#create it as a sub-dictionary. Otherwise, ignore
				aggregate_dict.setdefault(vendor_name, {})

				#If the item_id hasn't been seen before, create it as a
				#sub-dictionary. If it has, ignore this
				aggregate_dict[vendor_name].setdefault(item_id, {
					'cat_plus_subcat': cat_plus_subcat,
					'total_expertise': 0, 'total_weighted_score': 0})
				#Add the expertise for this row (scorer) to the total
				#expertise for this vendor and item
				aggregate_dict[vendor_name][item_id]['total_expertise'] += expertise
				#Add the expertise-weighted score for this item and
				#vendor to the total weighted score
				aggregate_dict[vendor_name][item_id]['total_weighted_score'] += weighted_score

	print(f"Saving results to aggregate file...")
	result_file = open('aggregate.py', 'w')
	result_file.write('aggregate_dict = ' + pprint.pformat(aggregate_dict) + '\n')
	result_file.close()
	print('Done.')

	return


def average_scores(vendors, nsv, num_of_vendors):
	"""Use the text file aggregate.py to determine average scores
	per vendor"""
	import os
	from aggregate import aggregate_dict as a_d

	#Create empty average scores file to be appended later
	result_file = open('average_scores.py', 'w')
	result_file.close()

	for vendor_name in a_d.keys():
		average_scores_dict = {}

		for item_id in a_d[vendor_name].keys():
			#Get data from the aggregate dictionaries
			total_weighted_score = a_d[vendor_name][item_id]['total_weighted_score']
			total_expertise = a_d[vendor_name][item_id]['total_expertise']
			cat_plus_subcat = a_d[vendor_name][item_id]['cat_plus_subcat']

			#calculate average score for each item
			if total_expertise != 0:
				average_score = total_weighted_score / total_expertise
			else:
				average_score = 0

			average_scores_dict.setdefault(item_id, {})
			#Store item info and average score in new average_scores_dict
			average_scores_dict[item_id]['cat_plus_subcat'] = cat_plus_subcat
			average_scores_dict[item_id]['average_score'] = average_score

		print(f"Saving {vendor_name} results to average scores file...")
		result_file = open('average_scores.py', 'a')
		result_file.write(f'{vendor_name} = ' + pprint.pformat(average_scores_dict) + '\n')
		result_file.close()
		print('Done.')


def open_weightings(weighting_name):
	"""Opens the weighting sheet and stores its contents in a Python
	dictionary."""

	while True:
		print(f'Opening {weighting_name}...')
		try:
			wb = openpyxl.load_workbook(f'{weighting_name}.xlsx', data_only=True)
		except FileNotFoundError:
			print(f"{weighting_name}.xlsx NOT FOUND! Please check the "
				"file name")
			input("Press 'enter' to re-enter the weighting sheet name "
				"and try again.")
			weighting_name = get_weighting_name()
			print()
			continue
		else:
			break

	sheet = wb['Item Weightings']

	weightings = {}

	for row in range(4, sheet.max_row + 1):
		item_id = sheet['A' + str(row)].value
		weight = sheet['H' + str(row)].value
		try:
			weight = float(0 if weight is None else weight)
		except ValueError:
			print(f"ERROR IN ROW {row}! SETTING VALUE TO 0. CHECK RESULTS")
			weight = 0
			continue

		#If the item_id hasn't been seen before, create it as a
		#sub-dictionary. If it has, ignore this
		weightings.setdefault(item_id, 0)
		#Add the weighting for this row (item_id)
		weightings[item_id] = weight

	print(f"Saving weightings to weightings file...")
	result_file = open('weightings.py', 'w')
	result_file.write('weightings = ' + pprint.pformat(weightings))
	result_file.close()
	print('Done.')

	return weighting_name


def final_scores(vendors, nsv, num_of_vendors):
	"""Use the text file aggregate.py to determine average scores
	per vendor"""
	import os, average_scores, weightings

	#Create empty final scores file to be appended later
	result_file = open('final_scores.py', 'w')
	result_file.close()

	for vendor in nsv:
		#Get vendor dict for current vendor
		v_d = getattr(average_scores, vendor)
		#Create empty average scores dict for vendor
		final_scores_dict = {}

		for item_id in v_d.keys():
			if item_id is not None:
				#Get data from the aggregate dictionaries
				cat_plus_subcat = v_d[item_id]['cat_plus_subcat']
				average_score = v_d[item_id]['average_score']
				weight = weightings.weightings[item_id]

				#calculate final score for each item
				final_score = average_score * weight

				final_scores_dict.setdefault(item_id, {})
				#Store item info and average score in new average_scores_dict
				final_scores_dict[item_id]['cat_plus_subcat'] = cat_plus_subcat
				final_scores_dict[item_id]['final_score'] = final_score

		print(f"Saving {vendor} results to final scores file...")
		result_file = open('final_scores.py', 'a')
		result_file.write(f'{vendor} = ' + pprint.pformat(final_scores_dict) + '\n')
		result_file.close()
		print('Done.')


def results(vendors, nsv, num_of_vendors):
	"""Use the text file aggregate.py to determine average scores
	per vendor"""
	import os, final_scores

	#Create blank results file to be appended later
	result_file = open('results.py', 'w')
	result_file.close()

	for vendor in nsv:
		#Get vendor dict for current vendor
		v_d = getattr(final_scores, vendor)
		#Create empty average scores dict for vendor
		results_dict = {}
		results_dict['Total'] = 0

		for item_id in v_d.keys():
			#Get data from the aggregate dictionaries
			cat_plus_subcat = v_d[item_id]['cat_plus_subcat']
			final_score = v_d[item_id]['final_score']

			#Make sure the subcategory for this item exists.
			#If it does, skip this
			results_dict.setdefault(cat_plus_subcat, {'subcat_score': 0})

			#Add item score to subcat score and total
			results_dict[cat_plus_subcat]['subcat_score'] += final_score
			results_dict['Total'] += final_score

			#Remove cat+subcat headings (scores are empty)
			if results_dict[cat_plus_subcat]['subcat_score'] == 0:
				del results_dict[cat_plus_subcat]

		print(f"Saving {vendor} results to results file...")
		result_file = open('results.py', 'a')
		result_file.write(f'{vendor} = ' + pprint.pformat(results_dict) + '\n')
		result_file.close()
		print('Done.')

	print('\nCOMPILATION COMPLETE. Open "results.xlsx" in the folder to view result.')


def create_excel_doc(vendors, nsv, weighting_name):
	"""Turn results module dictionaries into Excel spreadsheet"""
	import results

	wb = openpyxl.Workbook()
	sheet = wb['Sheet']

	static_headers = ['Cat+Subcat']
	headers = static_headers + vendors

	num_of_headers = len(headers)

	#Create header cells atop each column in row 3
	for i in range(1, num_of_headers + 1):
		column = get_column_letter(i)
		sheet[f'{column}3'] = headers[i-1]

	#Create Cat+Subcats from weighting sheet down first colum
	weight_wb = openpyxl.load_workbook(f'{weighting_name}.xlsx',
		data_only=True)
	cat_weights_sheet = weight_wb['Category Weightings']

	#Find last row of catweights table. Can't use .max_row because
	#Additional rows may exist beyond table
	for row in range(3, cat_weights_sheet.max_row + 1):
		if cat_weights_sheet[f'A{row}'].value == None or \
		cat_weights_sheet[f'A{row}'].value == 0:
			last_row = row - 1
			break

	#Add Cat+Subcats as row headers in the order from the weighting sheet
	for row in range(4, last_row + 1):
		sheet[f'A{row}'] = cat_weights_sheet[f'A{row}'].value
	#Add a "total" row
	row += 1
	sheet[f'A{row}'] = 'Total'

	#Get cat+subcat scores for each vendor
	for i in range(0, num_of_vendors):
		column = get_column_letter(i+2)
		v_d = getattr(results, nsv[i])
		#Have to remove 'total' k-v pair because it doesn't follow the
		#same k-v pattern as rest of dict
		total = v_d['Total']
		del v_d['Total']
		for row in range(4, last_row + 1):
			cat_plus_subcat = sheet[f'A{row}'].value
			subcat_score = v_d[cat_plus_subcat]['subcat_score']
			sheet[f'{column}{row}'] = subcat_score
		#Add total to bottom
		row += 1
		sheet[f'{column}{row}'] = total


	wb.save('results.xlsx')
	

base_name, num_of_scorecards, sheet_name, weighting_name = get_user_data()

vendors, nsv, num_of_vendors, base_name, sheet_name = get_vendors(
	base_name, sheet_name)

create_aggregate_file(base_name, num_of_scorecards, sheet_name,
	vendors, nsv, num_of_vendors)

average_scores(vendors, nsv, num_of_vendors)

weighting_name = open_weightings(weighting_name)

final_scores(vendors, nsv, num_of_vendors)

results(vendors, nsv, num_of_vendors)

create_excel_doc(vendors, nsv, weighting_name)